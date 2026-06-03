# -*- coding: utf-8 -*-
import uuid
import logging
from datetime import date, timedelta
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError, UserError

_logger = logging.getLogger(__name__)


class ResPartner(models.Model):
    """
    Estensione di res.partner per BuildingPay v15.

    Aggiunge:
    - Tipo contatto 'Condominio' (con icona edificio nel form backend)
    - Flag Amministratore BuildingPay
    - Codice Referrer (8 car.) + link di invito computato
    - Dati bancari: IBAN e banca tramite res.partner.bank nativo Odoo
    - Listino: usa property_product_pricelist nativo Odoo
    - Sezione portale 'Contratti' con due documenti:
        1. Accordo retrocessioni amministratore ED
           (placeholder: nome, CF, IBAN, banca, data)
        2. Accordo Condomini Aggregati ED
           (placeholder: nome, CF + Allegato A auto-compilato con condomini attivi)
    - Flag: accordo_retrocessioni_ed, accordo_condomini_aggregati_ed
    - Campi italiani: fiscalcode, pec_mail, codice_destinatario
    """
    _inherit = 'res.partner'

    # -------------------------------------------------------
    # Tipo contatto esteso con 'condominio'
    # -------------------------------------------------------
    type = fields.Selection(
        selection_add=[('condominio', 'Condominio')],
        ondelete={'condominio': 'set default'},
    )

    # -------------------------------------------------------
    # Flag e dati Amministratore
    # -------------------------------------------------------
    is_amministratore = fields.Boolean(
        string='Amministratore',
        default=False,
        tracking=True,
    )
    is_amministratore_validato = fields.Boolean(
        string='Amministratore validato',
        default=False,
        tracking=True,
        help=(
            'Se TRUE, l\'amministratore ha superato il controllo manuale '
            'dei dati inseriti e può operare sul portale (es. creare condomini). '
            'Impostato a FALSE automaticamente al momento della registrazione.'
        ),
    )
    referrer_code = fields.Char(
        string='Codice Referrer',
        copy=False,
        index=True,
        size=8,
        help=(
            'Codice univoco di 8 caratteri per il link di invito. '
            'Viene auto-generato quando si abilita il flag Amministratore. '
            'Può essere modificato manualmente o rigenerato con il pulsante.'
        ),
    )
    referral_url = fields.Char(
        string='Link di invito',
        compute='_compute_referral_url',
        store=False,
        help="URL completo che l'amministratore può inviare ai suoi clienti per registrarsi.",
    )
    referrer_id = fields.Many2one(
        comodel_name='res.partner',
        string='Referrer',
        ondelete='set null',
        tracking=True,
    )
    referred_ids = fields.One2many(
        comodel_name='res.partner',
        inverse_name='referrer_id',
        string='Amministratori invitati',
    )

    # -------------------------------------------------------
    # Valore retrocessione Amministratore (valuta aziendale)
    # -------------------------------------------------------
    company_currency_id = fields.Many2one(
        comodel_name='res.currency',
        string='Valuta aziendale',
        compute='_compute_company_currency_id',
        store=False,
    )
    valore_retrocessione_amministratore = fields.Monetary(
        string='Valore retrocessione amministratore',
        currency_field='company_currency_id',
        digits=(16, 2),
        default=0.20,
        tracking=True,
        help='Valore di retrocessione riconosciuto all\'amministratore, nella valuta aziendale.',
    )

    # -------------------------------------------------------
    # Flag e dati retrocessione Referrer
    # -------------------------------------------------------
    is_referrer = fields.Boolean(
        string='È un referrer',
        default=False,
        tracking=True,
        help='Diventa TRUE automaticamente quando viene generato il Codice Referrer.',
    )
    use_referrer_salesperson = fields.Boolean(
        string='Usare addetto vendite presente in anagrafica nel LEAD',
        default=False,
        tracking=True,
        help=(
            'TRUE: l\'addetto vendite del LEAD generato da questo referrer è '
            'l\'addetto vendite configurato in anagrafica (se impostato), '
            'altrimenti si usa il default nelle Configurazioni BuildingPay.\n'
            'FALSE: si usa sempre l\'addetto vendite di default delle Configurazioni BuildingPay.'
        ),
    )
    retrocessione_currency_id = fields.Many2one(
        comodel_name='res.currency',
        string='Valuta retrocessione',
        default=lambda self: self.env.company.currency_id,
    )
    valore_retrocessione = fields.Monetary(
        string='Valore retrocessione',
        currency_field='retrocessione_currency_id',
        default=0.08,
        tracking=True,
        help='Valore fisso di retrocessione nella valuta selezionata.',
    )
    contratto_retrocessione_standard = fields.Boolean(
        string='Contratto retrocessione standard',
        default=True,
        tracking=True,
        help=(
            'TRUE: sul portale viene mostrato e fatto scaricare il template '
            '"Accordo Retrocessioni" caricato nella Configurazione BuildingPay.\n'
            'FALSE: viene usato il file personalizzato caricato qui sotto (se presente). '
            'Se nessun file è caricato, la sezione "Accordo Retrocessioni" '
            'non viene mostrata sul portale.'
        ),
    )
    contratto_retrocessione_custom_file = fields.Binary(
        string='Contratto retrocessioni amministratore (personalizzato)',
        attachment=True,
        help='File .docx o .pdf da mostrare in sostituzione del template standard.',
    )
    contratto_retrocessione_custom_filename = fields.Char(
        string='Nome file contratto personalizzato',
    )

    # -------------------------------------------------------
    # Privacy
    # -------------------------------------------------------
    privacy_accepted = fields.Boolean(
        string='Privacy accettata',
        default=False,
        tracking=True,
        help='True se il contatto ha accettato la privacy policy al momento della registrazione.',
    )
    privacy_accepted_date = fields.Datetime(
        string='Data accettazione privacy',
        readonly=True,
    )

    # -------------------------------------------------------
    # CONTRATTO 1: Accordo retrocessioni amministratore ED
    # Placeholder nel template:
    #   [NOME AMMINISTRATORE]  → partner.name
    #   [CODICE FISCALE]       → partner.fiscalcode
    #   [IBAN]                 → primo res.partner.bank.acc_number
    #   [NOME BANCA]           → bank_id.name su res.partner.bank
    #   [DATA]                 → data odierna DD/MM/YYYY
    # -------------------------------------------------------
    accordo_retrocessioni_ed = fields.Boolean(
        string='Accordo retrocessioni amministratore ED',
        default=False,
        tracking=True,
        help='Attivato quando l\'amministratore carica l\'Accordo Retrocessioni firmato.',
    )
    accordo_retrocessioni_file = fields.Binary(
        string='File Accordo Retrocessioni',
        attachment=True,
    )
    accordo_retrocessioni_filename = fields.Char(
        string='Nome file accordo retrocessioni',
    )
    accordo_retrocessioni_upload_date = fields.Datetime(
        string='Data caricamento accordo retrocessioni',
        readonly=True,
    )

    # -------------------------------------------------------
    # CONTRATTO 2: Accordo Condomini Aggregati ED
    # Placeholder nel template:
    #   [NOME AMMINISTRATORE]  → partner.name
    #   [________]             → partner.fiscalcode
    #   [ALLEGATO_A]           → tabella condomini attivi (nome | indirizzo | IBAN)
    # -------------------------------------------------------
    accordo_condomini_aggregati_ed = fields.Boolean(
        string='Accordo condomini aggregati ED',
        default=False,
        tracking=True,
        help='Attivato quando l\'amministratore carica l\'Accordo Condomini Aggregati firmato.',
    )
    accordo_condomini_file = fields.Binary(
        string='File Accordo Condomini Aggregati',
        attachment=True,
    )
    accordo_condomini_filename = fields.Char(
        string='Nome file accordo condomini',
    )
    accordo_condomini_upload_date = fields.Datetime(
        string='Data caricamento accordo condomini',
        readonly=True,
    )

    # -------------------------------------------------------
    # Contratto / accordo firmato da Pes (caricato da backoffice)
    # Quando viene caricato, si invia automaticamente una mail
    # all'amministratore con il file in allegato.
    # -------------------------------------------------------
    contratto_amministratore_pes = fields.Binary(
        string='Contratto Amministratore firmato da Pes',
        attachment=True,
        help='Documento firmato da Pes. Al caricamento viene inviata una mail '
             'all\'amministratore con il file in allegato.',
    )
    contratto_amministratore_pes_filename = fields.Char(
        string='Nome file contratto Pes',
    )
    contratto_amministratore_pes_upload_date = fields.Datetime(
        string='Data caricamento contratto Pes',
        readonly=True,
        tracking=True,
    )
    contratto_amministratore_pes_upload_user_id = fields.Many2one(
        comodel_name='res.users',
        string='Caricato da',
        readonly=True,
        tracking=True,
    )
    accordo_condomini_pes = fields.Binary(
        string='Accordo Condomini Aggregati firmato da Pes',
        attachment=True,
        help='Documento firmato da Pes. Al caricamento viene inviata una mail '
             'all\'amministratore con il file in allegato.',
    )
    accordo_condomini_pes_filename = fields.Char(
        string='Nome file accordo condomini Pes',
    )
    accordo_condomini_pes_upload_date = fields.Datetime(
        string='Data caricamento accordo condomini Pes',
        readonly=True,
        tracking=True,
    )
    accordo_condomini_pes_upload_user_id = fields.Many2one(
        comodel_name='res.users',
        string='Caricato da',
        readonly=True,
        tracking=True,
    )

    # -------------------------------------------------------
    # Data archiviazione (per indirizzi di tipo condominio)
    # -------------------------------------------------------
    data_archiviazione = fields.Date(
        string='Data archiviazione',
        readonly=True,
    )
    dismesso_comunicato_delivery = fields.Boolean(
        string='Dismesso e già comunicato a Delivery',
        default=False,
        help='Impostato automaticamente dopo che il condominio dismesso è stato incluso '
             'nel report giornaliero. Viene azzerato se il condominio viene riattivato.',
    )

    # -------------------------------------------------------
    # Campi italiani (compatibili con l10n_it_edi)
    # fiscalcode e pec_mail sono già definiti da l10n_it_edi / base_address_extended
    # — non vengono ridefiniti qui per evitare conflitti.
    # -------------------------------------------------------
    codice_istat = fields.Char(
        string='Codice ISTAT',
        size=6,
        index=True,
        tracking=True,
        help=(
            'Codice ISTAT alfanumerico del comune (6 cifre con zero iniziale, es. "037006"). '
            'Valorizzato automaticamente al salvataggio in base al campo Città, '
            'ricercando nella tabella codici ISTAT importati. Modificabile manualmente.'
        ),
    )
    nome_promotore = fields.Char(
        string='Nome Promotore',
    )
    applicativo = fields.Char(
        string='Applicativo',
    )
    sistema_pagamento = fields.Char(
        string='Sistema di pagamento',
        size=10,
    )
    codice_destinatario = fields.Char(
        string='Codice Destinatario SDI',
        size=7,
    )
    electronic_invoice_subjected = fields.Boolean(
        string='Soggetto a fatturazione elettronica',
        default=False,
    )
    electronic_invoice_obliged_subject = fields.Boolean(
        string='Obbligo fatturazione elettronica',
        default=False,
    )

    # -------------------------------------------------------
    # Condizioni economiche per-amministratore
    # Valori di default letti dalla configurazione BuildingPay,
    # modificabili individualmente per ogni amministratore.
    # -------------------------------------------------------
    bp_costo_email = fields.Monetary(
        string='Costo email',
        currency_field='company_currency_id',
        default=lambda self: self._get_bp_config().costo_email,
        help='Costo per email inviata. Default: valore dalle Configurazioni BuildingPay.',
    )
    bp_costo_rendicontazione = fields.Monetary(
        string='Costo rendicontazione',
        currency_field='company_currency_id',
        default=lambda self: self._get_bp_config().costo_rendicontazione,
        help='Costo di rendicontazione. Default: valore dalle Configurazioni BuildingPay.',
    )
    bp_costo_whatsapp = fields.Monetary(
        string='Costo WhatsApp',
        currency_field='company_currency_id',
        default=lambda self: self._get_bp_config().costo_whatsapp,
        help='Costo per messaggio WhatsApp. Default: valore dalle Configurazioni BuildingPay.',
    )
    bp_quota_fissa = fields.Monetary(
        string='Quota fissa',
        currency_field='company_currency_id',
        default=lambda self: self._get_bp_config().quota_fissa,
        help='Quota fissa mensile. Default: valore dalle Configurazioni BuildingPay.',
    )
    bp_quota_fissa_sdd_product_id = fields.Many2one(
        comodel_name='product.product',
        string='Quota fissa SDD',
        domain="[('type', '=', 'service'), ('categ_id.name', '=', 'BuildingPay')]",
        default=lambda self: self._get_bp_config().quota_fissa_sdd_product_id,
        help='Prodotto SDD per questo amministratore. Default: prodotto dalle Configurazioni BuildingPay.',
    )

    # -------------------------------------------------------
    # Condomini figli (indirizzi di tipo condominio)
    # -------------------------------------------------------
    condominio_ids = fields.One2many(
        comodel_name='res.partner',
        inverse_name='parent_id',
        string='Condomini',
        domain=[('type', '=', 'condominio'), ('active', '=', True)],
    )
    condominio_count = fields.Integer(
        string='Numero Condomini',
        compute='_compute_condominio_count',
    )

    # -------------------------------------------------------
    # Compute
    # -------------------------------------------------------
    @api.depends('company_id')
    def _compute_company_currency_id(self):
        """Ritorna sempre la valuta dell'azienda corrente."""
        default_currency = self.env.company.currency_id
        for partner in self:
            partner.company_currency_id = (
                partner.company_id.currency_id or default_currency
            )

    @api.depends('child_ids', 'child_ids.type', 'child_ids.active')
    def _compute_condominio_count(self):
        for partner in self:
            partner.condominio_count = self.env['res.partner'].search_count([
                ('parent_id', '=', partner.id),
                ('type', '=', 'condominio'),
                ('active', '=', True),
            ])

    # -------------------------------------------------------
    # Lookup codice ISTAT dalla tabella importata
    # -------------------------------------------------------
    @api.model
    def _lookup_codice_istat_for_city(self, city_name):
        """Cerca il codice ISTAT alfanumerico nella tabella buildingpay.comune.istat
        confrontando denominazione_it con city_name (case-insensitive).
        Restituisce il codice alfanumerico (es. '037006') o False se non trovato."""
        if not city_name:
            return False
        istat = self.env['buildingpay.comune.istat'].sudo().search([
            ('denominazione_it', '=ilike', city_name.strip()),
        ], limit=1)
        return istat.cod_comune_alfanumerico or False

    # -------------------------------------------------------
    # Create / Write
    # -------------------------------------------------------
    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if 'city' in vals and 'codice_istat' not in vals:
                codice = self._lookup_codice_istat_for_city(vals.get('city', ''))
                if codice:
                    vals['codice_istat'] = codice
        return super().create(vals_list)

    def write(self, vals):
        # Se la città cambia e codice_istat non è già esplicitamente fornito,
        # auto-popola il codice ISTAT dalla tabella importata.
        if 'city' in vals and 'codice_istat' not in vals:
            city_name = vals.get('city', '')
            codice = self._lookup_codice_istat_for_city(city_name)
            vals = dict(vals)
            vals['codice_istat'] = codice or False

        # Aggiorna metadata upload per contratto_amministratore_pes
        # (data/ora e utente corrente) quando il campo viene modificato
        if 'contratto_amministratore_pes' in vals:
            vals = dict(vals)
            if vals.get('contratto_amministratore_pes'):
                vals['contratto_amministratore_pes_upload_date'] = fields.Datetime.now()
                vals['contratto_amministratore_pes_upload_user_id'] = self.env.user.id
            else:
                vals['contratto_amministratore_pes_upload_date'] = False
                vals['contratto_amministratore_pes_upload_user_id'] = False

        if 'accordo_condomini_pes' in vals:
            vals = dict(vals)
            if vals.get('accordo_condomini_pes'):
                vals['accordo_condomini_pes_upload_date'] = fields.Datetime.now()
                vals['accordo_condomini_pes_upload_user_id'] = self.env.user.id
            else:
                vals['accordo_condomini_pes_upload_date'] = False
                vals['accordo_condomini_pes_upload_user_id'] = False

        # Rileva la transizione is_amministratore_validato False → True
        # PRIMA di super().write() per confrontare con il valore attuale
        to_validate_email = []
        if vals.get('is_amministratore_validato') is True:
            for partner in self:
                if not partner.is_amministratore_validato and partner.is_amministratore:
                    to_validate_email.append(partner.id)

        # Rileva condomini che passano da inactive → active (disarchiviazione)
        # per azzerare il flag dismesso_comunicato_delivery
        to_reset_delivery = self.env['res.partner']
        if vals.get('active') is True:
            to_reset_delivery = self.with_context(active_test=False).filtered(
                lambda p: p.type == 'condominio' and not p.active
            )

        # Protegge la P.IVA dei figli di tipo condominio dalla sincronizzazione
        # automatica di Odoo (_fields_sync / _commercial_fields_sync).
        # Quando viene aggiornato il vat del partner commerciale (l'amministratore),
        # Odoo propaga il valore a tutti i figli — inclusi i condomini, che devono
        # invece mantenere la propria P.IVA indipendente.
        # Strategia: backup del vat di ogni figlio condominio prima del write,
        # ripristino dopo se il sync lo ha sovrascritto.
        condomini_vat_backup = {}
        if 'vat' in vals and not self.env.context.get('_bp_skip_vat_restore'):
            condomini = self.with_context(active_test=False).mapped('child_ids').filtered(
                lambda c: c.type == 'condominio'
            )
            condomini_vat_backup = {c.id: c.vat for c in condomini}

        result = super().write(vals)

        if to_reset_delivery:
            to_reset_delivery.write({'dismesso_comunicato_delivery': False})

        # Ripristina la P.IVA originale sui condomini modificati dal sync automatico
        if condomini_vat_backup:
            for child in self.browse(list(condomini_vat_backup.keys())):
                if child.vat != condomini_vat_backup[child.id]:
                    child.with_context(_bp_skip_vat_restore=True).write(
                        {'vat': condomini_vat_backup[child.id]}
                    )

        # Invia email di abilitazione per ogni amministratore appena validato
        for pid in to_validate_email:
            self.browse(pid)._send_amministratore_validato_email()

        # Trigger email quando viene caricato il contratto firmato da Pes
        if vals.get('contratto_amministratore_pes'):
            for partner in self:
                partner._send_pes_signed_email(
                    template_name='Invio contratto firmato',
                    file_data=partner.contratto_amministratore_pes,
                    filename=partner.contratto_amministratore_pes_filename or 'contratto_amministratore_pes.pdf',
                )
        if vals.get('accordo_condomini_pes'):
            for partner in self:
                partner._send_pes_signed_email(
                    template_name='Invio accordo condomini firmato',
                    file_data=partner.accordo_condomini_pes,
                    filename=partner.accordo_condomini_pes_filename or 'accordo_condomini_pes.pdf',
                )

        return result

    def _send_pes_signed_email(self, template_name, file_data, filename):
        """Invia la mail con il documento firmato da Pes in allegato.
        Cerca il template per nome (invariante tra versioni del modulo).
        Aggiunge il file come allegato alla mail prima dell'invio.
        """
        self.ensure_one()
        try:
            template = self.env['mail.template'].sudo().search([
                ('name', '=', template_name),
                ('model', '=', 'res.partner'),
            ], limit=1)
            if not template:
                _logger.warning(
                    'BuildingPay: template mail "%s" non trovato su res.partner.',
                    template_name)
                return

            # Crea l'allegato IR prima di generare la mail
            attachment = self.env['ir.attachment'].sudo().create({
                'name': filename,
                'datas': file_data,
                'res_model': 'res.partner',
                'res_id': self.id,
                'mimetype': 'application/pdf',
            })

            # Crea la mail dal template (force_send=False: la mandiamo dopo)
            mail_id = template.send_mail(self.id, force_send=False)
            mail = self.env['mail.mail'].sudo().browse(mail_id)
            if mail.exists():
                mail.attachment_ids = [(4, attachment.id)]
                mail.send()
                _logger.info(
                    'BuildingPay: mail "%s" inviata a %s (%s)',
                    template_name, self.name, self.email)
        except Exception as e:
            _logger.error(
                'BuildingPay: errore invio mail "%s" per partner %s: %s',
                template_name, self.id, e)

    @api.depends('referrer_code')
    def _compute_referral_url(self):
        """Calcola il link di invito: {domain}/web/signup?referrer={code}.
        Il dominio viene letto dal sito web "BuildingPay" (ricerca per nome,
        case-insensitive). Se non trovato o senza dominio configurato,
        fallback su web.base.url di sistema.
        """
        # Cerca il sito BuildingPay e usa il suo dominio
        bp_website = self.env['website'].sudo().search([
            ('name', 'ilike', 'BuildingPay'),
        ], limit=1)
        if bp_website and bp_website.domain:
            domain = bp_website.domain.strip().rstrip('/')
            # Aggiunge lo schema https:// se non presente
            if not domain.startswith('http'):
                domain = 'https://' + domain
        else:
            domain = self.env['ir.config_parameter'].sudo().get_param(
                'web.base.url', '').rstrip('/')

        for partner in self:
            if partner.referrer_code:
                partner.referral_url = '%s/web/signup?referrer=%s' % (
                    domain, partner.referrer_code)
            else:
                partner.referral_url = False

    def _send_amministratore_validato_email(self):
        """Invia la mail di abilitazione quando is_amministratore_validato viene
        impostato a TRUE. Cerca il template per nome (invariante tra versioni)."""
        self.ensure_one()
        try:
            if not self.email:
                _logger.warning(
                    'BuildingPay: mail validazione non inviata a %s: email mancante.',
                    self.name)
                return
            template = self.env['mail.template'].sudo().search([
                ('name', '=', 'Amministratore abilitato a BuildingPay'),
                ('model', '=', 'res.partner'),
            ], limit=1)
            if not template:
                _logger.warning(
                    'BuildingPay: template "Amministratore abilitato a BuildingPay" '
                    'non trovato.')
                return
            template.send_mail(self.id, force_send=True)
            _logger.info(
                'BuildingPay: mail abilitazione inviata a %s (%s)',
                self.name, self.email)
        except Exception as e:
            _logger.error(
                'BuildingPay: errore invio mail abilitazione per partner %s: %s',
                self.id, e)

    def action_generate_referrer_code(self):
        """Genera un nuovo codice referrer casuale di 8 caratteri.
        Imposta anche is_referrer = True.
        Ritorna False in modo che Odoo 17 ricarichi il form e mostri subito
        il nuovo codice senza dover fare refresh manuale della pagina.
        """
        self.ensure_one()
        self.write({
            'referrer_code': self._generate_referrer_code(),
            'is_referrer': True,
        })
        # False → Odoo 17 aggiorna automaticamente il record nel form corrente
        return False

    @api.model
    def _generate_referrer_code(self):
        """Genera un codice alfanumerico casuale di 8 caratteri (maiuscolo)."""
        return uuid.uuid4().hex[:8].upper()

    @api.model
    def _get_bp_config(self):
        """Restituisce la configurazione BuildingPay del sito web il cui nome
        contiene 'BuildingPay' (case-insensitive). Fallback alla prima config
        disponibile se non trovata."""
        bp_website = self.env['website'].sudo().search(
            [('name', 'ilike', 'BuildingPay')], limit=1)
        if bp_website:
            config = self.env['buildingpay_v36.config'].sudo().search(
                [('website_id', '=', bp_website.id)], limit=1)
            if config:
                return config
        return self.env['buildingpay_v36.config'].sudo().search([], limit=1)

    def action_reset_bp_defaults_from_config(self):
        """Reimposta le Condizioni Economiche con i valori di default dalla config BuildingPay.
        Funziona su uno o più record (form singolo e azione massiva sulla lista).
        """
        config = self._get_bp_config()
        if not config:
            raise UserError(_('Nessuna configurazione BuildingPay trovata.'))
        self.write({
            'bp_costo_email': config.costo_email,
            'bp_costo_rendicontazione': config.costo_rendicontazione,
            'bp_costo_whatsapp': config.costo_whatsapp,
            'bp_quota_fissa': config.quota_fissa,
            'bp_quota_fissa_sdd_product_id': config.quota_fissa_sdd_product_id.id or False,
        })
        return False

    # -------------------------------------------------------
    # Upload: Accordo Retrocessioni
    # -------------------------------------------------------
    def action_upload_retrocessioni(self, file_data, filename):
        """
        Chiamato dal portale quando l'utente carica l'Accordo Retrocessioni firmato.
        """
        self.ensure_one()
        self.write({
            'accordo_retrocessioni_file': file_data,
            'accordo_retrocessioni_filename': filename,
            'accordo_retrocessioni_ed': True,
            'accordo_retrocessioni_upload_date': fields.Datetime.now(),
        })

    # -------------------------------------------------------
    # Upload: Accordo Condomini Aggregati
    # -------------------------------------------------------
    def action_upload_accordo_condomini(self, file_data, filename):
        """
        Chiamato dal portale quando l'utente carica l'Accordo Condomini Aggregati firmato.
        Attiva il flag accordo_condomini_aggregati_ed e crea eventuale attività automatica.
        """
        self.ensure_one()
        self.write({
            'accordo_condomini_file': file_data,
            'accordo_condomini_filename': filename,
            'accordo_condomini_aggregati_ed': True,
            'accordo_condomini_upload_date': fields.Datetime.now(),
        })
        self._create_contratto_activity()

    def _create_contratto_activity(self):
        """Crea un ToDo per ogni assegnatario configurato quando l'amministratore
        carica l'Accordo Condomini Aggregati firmato dal portale."""
        self.ensure_one()
        config = self.env['buildingpay_v36.config'].get_config_for_website()
        if not config or not config.create_activity_on_contract:
            return

        # Raccoglie tutti gli assegnatari configurati (fino a 4)
        responsabili = [
            config.activity_responsible_id,
            config.activity_responsible_2_id,
            config.activity_responsible_3_id,
            config.activity_responsible_4_id,
        ]
        responsabili = [r for r in responsabili if r]
        if not responsabili:
            _logger.warning(
                'BuildingPay: attività accordo condomini non creata: '
                'nessun assegnatario configurato')
            return

        deadline = date.today() + timedelta(days=config.activity_days or 5)
        activity_type = self.env.ref(
            'mail.mail_activity_data_todo', raise_if_not_found=False)

        summary = _('Controllare il contratto Accordo Condomini Aggregati '
                    "caricato dall'amministratore")
        for user in responsabili:
            self.activity_schedule(
                activity_type_id=activity_type.id if activity_type else False,
                summary=summary,
                date_deadline=deadline,
                user_id=user.id,
            )
            _logger.info(
                'BuildingPay: ToDo "Accordo condomini" creato per %s (assegnatario: %s)',
                self.name, user.name)

    # -------------------------------------------------------
    # Archiviazione condominio
    # -------------------------------------------------------
    def action_archive_condominio(self):
        """Archivia un indirizzo di tipo condominio."""
        self.ensure_one()
        if self.type != 'condominio':
            raise UserError(_(
                'Solo gli indirizzi di tipo "Condominio" possono essere archiviati.'))
        self.write({
            'data_archiviazione': fields.Date.today(),
            'active': False,
        })
        self._send_condominio_dismesso_email()

    def _send_condominio_dismesso_email(self):
        """Genera Excel condominio dismesso e lo invia via email."""
        self.ensure_one()
        config = self.env['buildingpay_v36.config'].get_config_for_website()
        if not config or not config.condomini_dismessi_email:
            return

        try:
            import openpyxl
            from io import BytesIO
            import base64

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Condominio Dismesso'
            headers = [
                'ID Esterno Amministratore', 'Nome Amministratore',
                'ID Esterno Condominio', 'Nome Condominio',
                'Indirizzo Completo', 'Dismesso',
            ]
            for col, h in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=h).font = \
                    openpyxl.styles.Font(bold=True)

            admin = self.parent_id
            admin_ext = self.env['ir.model.data'].search([
                ('model', '=', 'res.partner'),
                ('res_id', '=', admin.id if admin else 0),
            ], limit=1)
            cond_ext = self.env['ir.model.data'].search([
                ('model', '=', 'res.partner'), ('res_id', '=', self.id),
            ], limit=1)
            address = ' '.join(filter(None, [
                self.street, self.zip, self.city,
                self.state_id.name if self.state_id else '',
                self.country_id.name if self.country_id else '',
            ]))
            row = [
                admin_ext.complete_name if admin_ext else '',
                admin.name if admin else '',
                cond_ext.complete_name if cond_ext else '',
                self.name or '', address, True,
            ]
            for col, v in enumerate(row, 1):
                ws.cell(row=2, column=col, value=v)

            out = BytesIO()
            wb.save(out)
            recipients = [
                e.strip() for e in config.condomini_dismessi_email.split(',')
                if e.strip()
            ]
            if recipients:
                self.env['mail.mail'].sudo().create({
                    'subject': _('Condominio dismesso: %s') % self.name,
                    'body_html': _(
                        '<p>Il condominio <b>%s</b> è stato dismesso '
                        'in data %s.</p>') % (self.name, fields.Date.today()),
                    'email_to': ','.join(recipients),
                    'attachment_ids': [(0, 0, {
                        'name': 'condominio_dismesso_%s.xlsx' % self.name,
                        'datas': base64.b64encode(out.getvalue()),
                        'mimetype': ('application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet'),
                    })],
                }).send()
        except Exception as e:
            _logger.error('BuildingPay v30: errore email condominio dismesso: %s', e)

    # -------------------------------------------------------
    # Azione pianificata: report giornaliero condomini attivi
    # Formato output: "Dati Enti Aggregati_PagoPa"
    #
    # Struttura file (2 righe di intestazione + dati da riga 3):
    #   Riga 1 — gruppi: A1:C1 "Identificativi" | D1:L1 "Ente" |
    #                    M1 "Piattaforma pagoPA" | N1:O1 "Dati Amministratore" | P1 "Stato"
    #   Riga 2 — colonne:
    #     A  ID Esterno Amministratore
    #     B  Nome Amministratore
    #     C  ID Esterno Condominio
    #     D  Ragione Sociale
    #     E  PEC
    #     F  Codice Fiscale
    #     G  P.IVA
    #     H  Sede legale - Indirizzo
    #     I  Sede legale - Città
    #     J  Sede legale - Provincia (sigla)
    #     K  CAP
    #     L  Codice Istat
    #     M  IBAN
    #     N  Applicativo (dall'amministratore)
    #     O  Sistema di Pagamento (dall'amministratore)
    # -------------------------------------------------------
    @api.model
    def action_send_daily_condomini_report(self):
        """Genera e invia il report Excel giornaliero dei condomini (attivi e dismessi
        non ancora comunicati) nel formato 'Dati Enti Aggregati_PagoPa'.
        Dopo l'invio, i condomini dismessi inclusi vengono marcati come
        dismesso_comunicato_delivery=True e non appariranno più nei report successivi.
        PEC: usa la PEC del condominio, con fallback sulla PEC dell'amministratore.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            import base64

            # Includi attivi + dismessi non ancora comunicati a Delivery
            condominii = self.with_context(active_test=False).search([
                ('type', '=', 'condominio'),
                ('parent_id', '!=', False),
                ('parent_id.is_amministratore', '=', True),
                '|',
                ('active', '=', True),
                ('dismesso_comunicato_delivery', '=', False),
            ])
            if not condominii:
                _logger.info('BuildingPay: nessun condominio trovato.')
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Enti Aggregati'

            # ── Stili ─────────────────────────────────────────
            bold_font       = Font(bold=True)
            white_bold_font = Font(bold=True, color='FFFFFF')
            fill_ente       = PatternFill('solid', fgColor='1F4E79')   # blu scuro
            fill_pagopa     = PatternFill('solid', fgColor='2E75B6')   # blu chiaro
            fill_ids        = PatternFill('solid', fgColor='D6E4F0')   # azzurro tenue
            fill_admin      = PatternFill('solid', fgColor='E2EFDA')   # verde tenue
            center          = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin_border     = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'),  bottom=Side(style='thin'),
            )

            # ── Riga 1: intestazioni di gruppo ────────────────
            # A1:C1 → area identificativi (sfondo azzurro tenue)
            # D1:L1 → "Ente" mergiate (blu scuro)
            # M1    → "Piattaforma pagoPA" (blu chiaro)
            # N1:O1 → "Dati Amministratore" (verde tenue)
            # P1    → "Stato" (grigio)
            fill_stato = PatternFill('solid', fgColor='E8E8E8')  # grigio chiaro

            ws.merge_cells('A1:C1')
            ws.merge_cells('D1:L1')
            ws.merge_cells('M1:N1')
            ws.merge_cells('O1:P1')
            ws['A1'] = 'Identificativi'
            ws['D1'] = 'Ente'
            ws['M1'] = 'Piattaforma pagoPA'
            ws['O1'] = 'Dati Amministratore'
            ws['Q1'] = 'Stato'

            for col_letter in ('A', 'B', 'C'):
                c = ws[col_letter + '1']
                c.fill = fill_ids
                c.font = bold_font
                c.alignment = center
                c.border = thin_border
            for col_letter in ('D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'):
                c = ws[col_letter + '1']
                c.fill = fill_ente
                c.font = white_bold_font
                c.alignment = center
                c.border = thin_border
            for col_letter in ('M', 'N'):
                c = ws[col_letter + '1']
                c.fill = fill_pagopa
                c.font = white_bold_font
                c.alignment = center
                c.border = thin_border
            for col_letter in ('O', 'P'):
                c = ws[col_letter + '1']
                c.fill = fill_admin
                c.font = bold_font
                c.alignment = center
                c.border = thin_border
            c = ws['Q1']
            c.fill = fill_stato
            c.font = bold_font
            c.alignment = center
            c.border = thin_border

            ws.row_dimensions[1].height = 22

            # ── Riga 2: intestazioni colonne ──────────────────
            col2_headers = [
                'ID Esterno Amministratore',       # A
                'Nome Amministratore',             # B
                'ID Esterno Condominio',           # C
                'Ragione Sociale',                 # D
                'PEC',                             # E
                'Codice Fiscale',                  # F
                'P.IVA',                           # G
                'Sede legale - Indirizzo',         # H
                'Sede legale - Città',             # I
                'Sede legale - Provincia (sigla)', # J
                'CAP',                             # K
                'Codice Istat',                    # L
                'IBAN',                            # M
                'IBAN secondario',                 # N
                'Applicativo',                     # O
                'Sistema di Pagamento',            # P
                'Stato',                           # Q
            ]
            for col_idx, header in enumerate(col2_headers, 1):
                c = ws.cell(row=2, column=col_idx, value=header)
                c.font = bold_font
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.border = thin_border
            ws.row_dimensions[2].height = 30

            # ── Larghezze colonne ─────────────────────────────
            col_widths = [30, 30, 30, 30, 30, 18, 18, 36, 22, 12, 8, 14, 32, 32, 24, 20, 12]
            for col_idx, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # ── Recupero ID esterni in batch ──────────────────
            all_admins    = condominii.mapped('parent_id')
            admin_ext_map = all_admins.get_external_id()
            cond_ext_map  = condominii.get_external_id()

            def _ext_id(partner, ext_map):
                xid = ext_map.get(partner.id, '')
                if not xid:
                    xid = '__export__.res_partner_%d' % partner.id
                return xid

            # ── Righe dati (partono da riga 3) ────────────────
            fill_dismesso = PatternFill('solid', fgColor='FDECEA')  # rosso tenue per dismessi

            for row_idx, condo in enumerate(condominii, 3):
                admin = condo.parent_id
                banks = self.with_context(active_test=False).env['res.partner.bank'].search([
                    ('partner_id', '=', condo.id),
                ], order='id asc', limit=2)
                bank  = banks[0] if len(banks) >= 1 else self.env['res.partner.bank']
                bank2 = banks[1] if len(banks) >= 2 else self.env['res.partner.bank']
                # PEC: usa la PEC del condominio se presente, altrimenti quella dell'amministratore
                pec_value = condo.pec_mail or admin.pec_mail or ''
                stato = 'attivo' if condo.active else 'dismesso'
                row_data = [
                    _ext_id(admin, admin_ext_map),                          # A
                    admin.name or '',                                        # B
                    _ext_id(condo, cond_ext_map),                           # C
                    condo.name or '',                                        # D
                    pec_value,                                               # E
                    condo.fiscalcode or '',                                  # F
                    condo.vat or '',                                         # G
                    condo.street or '',                                      # H
                    condo.city or '',                                        # I
                    condo.state_id.code if condo.state_id else '',          # J
                    condo.zip or '',                                         # K
                    condo.codice_istat or '',                                # L
                    bank.acc_number if bank else '',                         # M
                    bank2.acc_number if bank2 else '',                       # N
                    admin.applicativo or '',                                 # O
                    admin.sistema_pagamento or '',                           # P
                    stato,                                                   # Q
                ]
                for col_idx, value in enumerate(row_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=value)
                    c.border = thin_border
                    # Sfondo rosso tenue per le righe dei condomini dismessi
                    if not condo.active:
                        c.fill = fill_dismesso

            # ── Serializzazione e invio ───────────────────────
            out = BytesIO()
            wb.save(out)
            excel_data = base64.b64encode(out.getvalue())

            configs = self.env['buildingpay_v36.config'].search([
                ('condomini_attivati_email', '!=', False),
            ])
            all_recipients = set()
            for cfg in configs:
                for email in cfg.condomini_attivati_email.split(','):
                    if email.strip():
                        all_recipients.add(email.strip())

            if all_recipients:
                today_str = fields.Date.today().strftime('%Y-%m-%d')
                self.env['mail.mail'].sudo().create({
                    'subject': _('Dati Enti Aggregati PagoPa - %s') % today_str,
                    'body_html': _(
                        '<p>File Dati Enti Aggregati PagoPa del %s.</p>'
                    ) % today_str,
                    'email_to': ','.join(all_recipients),
                    'attachment_ids': [(0, 0, {
                        'name': 'Dati Enti Aggregati_PagoPa_%s.xlsx' % today_str,
                        'datas': excel_data,
                        'mimetype': ('application/vnd.openxmlformats-officedocument'
                                     '.spreadsheetml.sheet'),
                    })],
                }).send()

                # Marca i dismessi inclusi nel report: non verranno più inviati
                dismessi = condominii.filtered(lambda c: not c.active)
                if dismessi:
                    dismessi.sudo().write({'dismesso_comunicato_delivery': True})
                    _logger.info(
                        'BuildingPay: %d condominio/i dismesso/i marcato/i come '
                        'già comunicati a Delivery.', len(dismessi))
        except Exception as e:
            _logger.error('BuildingPay: errore report Dati Enti Aggregati: %s', e)
